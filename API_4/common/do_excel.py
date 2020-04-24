# -*- coding: utf-8 -*-
# @Time : 2020/4/7 0:22
# @Author : Bella
# @Email : 1171208366@qq.com
# @FileName : do_excel.py


class Case:
    """
    测试用例类，每个测试用例，实际上就是它的一个实例
    """""
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql =None



#完成excel的读和写

import openpyxl
from API_4.common import http_request

class DoExcel:
    def __init__(self,file_name,sheet_name): #读哪个excel，哪个excel中的sheet
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet = self.workbook[sheet_name]

    def get_cases(self):
        max_row = self.sheet.max_row

        cases = []
        for r in range(2,max_row+1):
            # case = {}
            # cases['case_id'] = self.sheet.cell(row=r,column=1)
            # cases['title'] = self.sheet.cell(row=r,column=2)

            case = Case() #实例
            case.case_id = self.sheet.cell(row=r,column=1).value
            case.title = self.sheet.cell(row=r,column=2).value
            case.url = self.sheet.cell(row=r,column=3).value
            case.data = self.sheet.cell(row=r,column=4).value
            case.method = self.sheet.cell(row=r,column=5).value
            case.expected = self.sheet.cell(row=r,column=6).value
            case.sql = self.sheet.cell(row=r,column=9).value

            cases.append(case)   #case变成了Case的实例了

        return cases #返回case列表

    def write_result(self,row,actual,result):
        sheet = self.workbook[self.sheet_name]
        sheet.cell(row,7).value = actual
        sheet.cell(row,8).value = result
        self.workbook.save(filename=self.file_name)
        self.workbook.close()


if __name__ == '__main__':
    from API_4.common import constants
    do_excel = DoExcel(constants.case_file,sheet_name='login')
    cases = do_excel.get_cases()
    http_request = http_request.Http_Request()

    for case in cases:
        # print(case.case_id)
        # print(case.method)
        # print(case.data)
        print(case.__dict__)   #每一个类都有这样一个属性__dict__  私有属性 这个属性实际上会返回所有的属性的字典
        resp = http_request.request(case.method,case.url,case.data)
        actual = resp.text #相应文本
        resp_dict = resp.json() #返回字典
        if case.expected == actual: #判断期望结果是否与实际结果一致
            do_excel.write_result(case.case_id+1,actual,'PASS')
        else:
            do_excel.write_result(case.case_id+1,actual,'FAILE')






